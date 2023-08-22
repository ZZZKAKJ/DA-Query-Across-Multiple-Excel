import os
import pyodbc
from openpyxl import load_workbook

def query_excel_data():
    # Declare variables
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "AllData")
    start_date = "2023-01"  # Modify accordingly
    end_date = "2023-12"  # Modify accordingly
    sum_quantity = 0

    # Set up database connection
    connection_str = (
        r"Driver={Microsoft Excel Driver (*.xls, *.xlsx, *.xlsm, *.xlsb)};"
        f"DBQ={os.path.join(path, 'your_file.xlsx')};"
    )
    connection = pyodbc.connect(connection_str)
    cursor = connection.cursor()

    # Get list of Excel files in the directory
    excel_files = [file for file in os.listdir(path) if file.endswith(".xlsx")]

    for excel_file in excel_files:
        # Check if the file falls within the specified date range
        if start_date <= excel_file[:7] <= end_date:
            print(os.path.join(path, excel_file))

            # Open Excel workbook using openpyxl
            wb = load_workbook(os.path.join(path, excel_file), data_only=True)
            sheet = wb.active

            # Execute SQL-like queries using openpyxl
            for row in sheet.iter_rows(min_row=2, values_only=True):
                sell_date = row[1]  # Assuming SellDate is in the second column
                if start_date <= sell_date <= end_date:
                    sum_quantity += row[5]  # Assuming Number is in the sixth column

            # Clean up Excel workbook
            wb.close()

    # Update the total sum (you need to handle the output)
    print("Total Sum:", sum_quantity)

    # Close database connection
    cursor.close()
    connection.close()

# Run the query function
query_excel_data()
